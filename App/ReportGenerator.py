import pandas as pd
import os
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

class ReportGenerator:
    def __init__(self, output_dir="reports"):
        """Initialize the report generator with an output directory."""
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_report(self, trade_data: list, risk_assessment: list, forecast_results: list, filename=None):
        """Generate an Excel report summarizing trade decisions, risk assessments, and forecasts."""
        try:
            if filename is None:
                filename = f"report_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            file_path = os.path.join(self.output_dir, filename)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Write Data to Separate Sheets
                trade_df = pd.DataFrame(trade_data)
                risk_df = pd.DataFrame(risk_assessment)
                forecast_df = pd.DataFrame(forecast_results)
                
                trade_df.to_excel(writer, sheet_name="Trade Decisions", index=False)
                risk_df.to_excel(writer, sheet_name="Risk Assessment", index=False)
                forecast_df.to_excel(writer, sheet_name="Forecast Results", index=False)

            # Apply Formatting and Charts
            self.format_report(file_path)
            self.add_charts(file_path, trade_df)

            return file_path
        except Exception as e:
            print(f"Error generating report: {e}")
            return None

    def format_report(self, file_path):
        """Apply formatting (bold headers, column widths, and conditional formatting)."""
        try:
            wb = load_workbook(file_path)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                
                # Format Headers
                for col_num, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
                    col_letter = get_column_letter(col_num)
                    ws[f"{col_letter}1"].font = Font(bold=True)
                    ws.column_dimensions[col_letter].width = 15

                # Conditional Formatting for RSI and MACD
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if "RSI" in ws.cell(row=1, column=cell.column).value and cell.value:
                            if cell.value > 70:
                                cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                            elif cell.value < 30:
                                cell.fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
                        elif "MACD_Hist" in ws.cell(row=1, column=cell.column).value and cell.value:
                            if cell.value > 0:
                                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                            elif cell.value < 0:
                                cell.fill = PatternFill(start_color="9999FF", end_color="9999FF", fill_type="solid")
            wb.save(file_path)
        except Exception as e:
            print(f"Error formatting report: {e}")

       
    
    def add_charts(self, file_path, trade_df):
        """Generate and embed charts in the report."""
        if trade_df.empty:
            return

        # Generate Price vs RSI Chart
        fig, ax1 = plt.subplots(figsize=(8, 4))
        ax1.set_xlabel("Date")
        ax1.set_ylabel("Close Price", color="blue")
        ax1.plot(trade_df['Date'], trade_df['Close'], label="Close Price", color="blue")
        
        ax2 = ax1.twinx()
        ax2.set_ylabel("RSI", color="red")
        ax2.plot(trade_df['Date'], trade_df['RSI'], label="RSI", color="red", linestyle="dashed")
        
        plt.title("Price vs RSI")
        plt.legend()
        plt.grid()
        plt.tight_layout()

        chart_path = os.path.join(self.output_dir, "price_vs_rsi.png")
        plt.savefig(chart_path)
        plt.close()

        # Embed Chart in Excel
        wb = load_workbook(file_path)
        ws = wb.create_sheet("Charts")
        img = Image(chart_path)
        ws.add_image(img, "A1")
        wb.save(file_path)

# Example usage:
# report_gen = ReportGenerator()
# report_file = report_gen.generate_report(trade_data, risk_assessment, forecast_results)
# print(f"Report saved at: {report_file}")
